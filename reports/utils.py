import io
import os
from datetime import datetime
from decimal import Decimal
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
import csv


class ReportExporter:
    """Utility class for exporting reports in different formats"""
    
    def __init__(self, report_data, report_type, metadata=None):
        self.report_data = report_data
        self.report_type = report_type
        self.metadata = metadata or {}
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def export_to_pdf(self, include_charts=True):
        """Export report to PDF format"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        title = f"{self.report_type.replace('_', ' ').title()} Report"
        story.append(Paragraph(title, title_style))
        
        # Metadata
        if self.metadata:
            meta_style = styles['Normal']
            for key, value in self.metadata.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", meta_style))
            story.append(Spacer(1, 20))
        
        # Report data
        if self.report_type == 'working_hours':
            self._add_working_hours_to_pdf(story, styles)
        elif self.report_type == 'overtime':
            self._add_overtime_to_pdf(story, styles)
        elif self.report_type == 'leave':
            self._add_leave_to_pdf(story, styles)
        elif self.report_type == 'payroll_summary':
            self._add_payroll_summary_to_pdf(story, styles)
        elif self.report_type == 'employee_performance':
            self._add_performance_to_pdf(story, styles)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_to_excel(self, include_charts=True):
        """Export report to Excel format"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"{self.report_type.replace('_', ' ').title()} Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = f"{self.report_type.replace('_', ' ').title()} Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Metadata
        row = 3
        if self.metadata:
            for key, value in self.metadata.items():
                worksheet[f'A{row}'] = f"{key}:"
                worksheet[f'B{row}'] = str(value)
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
            row += 1
        
        # Data
        if self.report_type == 'working_hours':
            self._add_working_hours_to_excel(worksheet, row, header_font, header_fill, header_alignment)
        elif self.report_type == 'overtime':
            self._add_overtime_to_excel(worksheet, row, header_font, header_fill, header_alignment)
        elif self.report_type == 'leave':
            self._add_leave_to_excel(worksheet, row, header_font, header_fill, header_alignment)
        elif self.report_type == 'payroll_summary':
            self._add_payroll_summary_to_excel(worksheet, row, header_font, header_fill, header_alignment)
        elif self.report_type == 'employee_performance':
            self._add_performance_to_excel(worksheet, row, header_font, header_fill, header_alignment)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_to_csv(self):
        """Export report to CSV format"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Write metadata
        if self.metadata:
            for key, value in self.metadata.items():
                writer.writerow([key, value])
            writer.writerow([])  # Empty row
        
        # Write data based on report type
        if self.report_type == 'working_hours':
            self._add_working_hours_to_csv(writer)
        elif self.report_type == 'overtime':
            self._add_overtime_to_csv(writer)
        elif self.report_type == 'leave':
            self._add_leave_to_csv(writer)
        elif self.report_type == 'payroll_summary':
            self._add_payroll_summary_to_csv(writer)
        elif self.report_type == 'employee_performance':
            self._add_performance_to_csv(writer)
        
        buffer.seek(0)
        return buffer
    
    def _add_working_hours_to_pdf(self, story, styles):
        """Add working hours data to PDF"""
        if not self.report_data:
            story.append(Paragraph("No data available for the selected period.", styles['Normal']))
            return
        
        # Create table data
        headers = [
            'Employee ID', 'Name', 'Department', 'Present Days',
            'Total Hours', 'Overtime Hours', 'Attendance Rate'
        ]
        
        data = [headers]
        for employee in self.report_data:
            data.append([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                str(employee['present_days']),
                str(employee['total_hours']),
                str(employee['overtime_hours']),
                f"{employee['attendance_rate']}%"
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_working_hours_to_excel(self, worksheet, start_row, header_font, header_fill, header_alignment):
        """Add working hours data to Excel"""
        headers = [
            'Employee ID', 'Name', 'Department', 'Present Days',
            'Total Hours', 'Overtime Hours', 'Attendance Rate'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, employee in enumerate(self.report_data, start_row + 1):
            worksheet.cell(row=row_idx, column=1, value=employee['employee_id'])
            worksheet.cell(row=row_idx, column=2, value=employee['employee_name'])
            worksheet.cell(row=row_idx, column=3, value=employee['department'])
            worksheet.cell(row=row_idx, column=4, value=employee['present_days'])
            worksheet.cell(row=row_idx, column=5, value=float(employee['total_hours']))
            worksheet.cell(row=row_idx, column=6, value=float(employee['overtime_hours']))
            worksheet.cell(row=row_idx, column=7, value=f"{employee['attendance_rate']}%")
    
    def _add_working_hours_to_csv(self, writer):
        """Add working hours data to CSV"""
        headers = [
            'Employee ID', 'Name', 'Department', 'Present Days',
            'Total Hours', 'Overtime Hours', 'Attendance Rate'
        ]
        writer.writerow(headers)
        
        for employee in self.report_data:
            writer.writerow([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                employee['present_days'],
                employee['total_hours'],
                employee['overtime_hours'],
                f"{employee['attendance_rate']}%"
            ])
    
    def _add_overtime_to_pdf(self, story, styles):
        """Add overtime data to PDF"""
        if not self.report_data:
            story.append(Paragraph("No overtime data available for the selected period.", styles['Normal']))
            return
        
        headers = ['Employee ID', 'Name', 'Department', 'Overtime Hours', 'Overtime Days', 'Overtime Amount']
        data = [headers]
        
        for employee in self.report_data:
            data.append([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                str(employee['total_overtime_hours']),
                str(employee['overtime_days']),
                f"${employee['overtime_amount']}"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_overtime_to_excel(self, worksheet, start_row, header_font, header_fill, header_alignment):
        """Add overtime data to Excel"""
        headers = ['Employee ID', 'Name', 'Department', 'Overtime Hours', 'Overtime Days', 'Overtime Amount']
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, employee in enumerate(self.report_data, start_row + 1):
            worksheet.cell(row=row_idx, column=1, value=employee['employee_id'])
            worksheet.cell(row=row_idx, column=2, value=employee['employee_name'])
            worksheet.cell(row=row_idx, column=3, value=employee['department'])
            worksheet.cell(row=row_idx, column=4, value=float(employee['total_overtime_hours']))
            worksheet.cell(row=row_idx, column=5, value=employee['overtime_days'])
            worksheet.cell(row=row_idx, column=6, value=float(employee['overtime_amount']))
    
    def _add_overtime_to_csv(self, writer):
        """Add overtime data to CSV"""
        headers = ['Employee ID', 'Name', 'Department', 'Overtime Hours', 'Overtime Days', 'Overtime Amount']
        writer.writerow(headers)
        
        for employee in self.report_data:
            writer.writerow([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                employee['total_overtime_hours'],
                employee['overtime_days'],
                employee['overtime_amount']
            ])
    
    def _add_leave_to_pdf(self, story, styles):
        """Add leave data to PDF"""
        if not self.report_data:
            story.append(Paragraph("No leave data available for the selected period.", styles['Normal']))
            return
        
        headers = ['Employee ID', 'Name', 'Department', 'Leaves Taken', 'Annual Balance', 'Sick Balance']
        data = [headers]
        
        for employee in self.report_data:
            data.append([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                str(employee['total_leaves_taken']),
                str(employee['annual_leave_balance']),
                str(employee['sick_leave_balance'])
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_leave_to_excel(self, worksheet, start_row, header_font, header_fill, header_alignment):
        """Add leave data to Excel"""
        headers = ['Employee ID', 'Name', 'Department', 'Leaves Taken', 'Annual Balance', 'Sick Balance']
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, employee in enumerate(self.report_data, start_row + 1):
            worksheet.cell(row=row_idx, column=1, value=employee['employee_id'])
            worksheet.cell(row=row_idx, column=2, value=employee['employee_name'])
            worksheet.cell(row=row_idx, column=3, value=employee['department'])
            worksheet.cell(row=row_idx, column=4, value=employee['total_leaves_taken'])
            worksheet.cell(row=row_idx, column=5, value=employee['annual_leave_balance'])
            worksheet.cell(row=row_idx, column=6, value=employee['sick_leave_balance'])
    
    def _add_leave_to_csv(self, writer):
        """Add leave data to CSV"""
        headers = ['Employee ID', 'Name', 'Department', 'Leaves Taken', 'Annual Balance', 'Sick Balance']
        writer.writerow(headers)
        
        for employee in self.report_data:
            writer.writerow([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                employee['total_leaves_taken'],
                employee['annual_leave_balance'],
                employee['sick_leave_balance']
            ])
    
    def _add_payroll_summary_to_pdf(self, story, styles):
        """Add payroll summary to PDF"""
        data = self.report_data
        
        # Summary section
        story.append(Paragraph("Payroll Summary", styles['Heading2']))
        summary_data = [
            ['Total Employees', str(data['total_employees'])],
            ['Total Gross Salary', f"${data['total_gross_salary']}"],
            ['Total Net Salary', f"${data['total_net_salary']}"],
            ['Total Deductions', f"${data['total_deductions']}"],
            ['Total Bonuses', f"${data['total_bonuses']}"],
            ['Average Salary', f"${data['average_net_salary']}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Department breakdown
        if data.get('department_breakdown'):
            story.append(Paragraph("Department Breakdown", styles['Heading3']))
            dept_headers = ['Department', 'Employees', 'Total Gross', 'Average Salary']
            dept_data = [dept_headers]
            
            for dept in data['department_breakdown']:
                dept_data.append([
                    dept.get('department', 'N/A'),
                    str(dept.get('employee_count', 0)),
                    f"${dept.get('total_gross', 0)}",
                    f"${dept.get('avg_salary', 0)}"
                ])
            
            dept_table = Table(dept_data)
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(dept_table)
    
    def _add_payroll_summary_to_excel(self, worksheet, start_row, header_font, header_fill, header_alignment):
        """Add payroll summary to Excel"""
        data = self.report_data
        
        # Summary section
        worksheet.cell(row=start_row, column=1, value="Payroll Summary").font = Font(bold=True, size=14)
        start_row += 2
        
        summary_items = [
            ('Total Employees', data['total_employees']),
            ('Total Gross Salary', float(data['total_gross_salary'])),
            ('Total Net Salary', float(data['total_net_salary'])),
            ('Total Deductions', float(data['total_deductions'])),
            ('Total Bonuses', float(data['total_bonuses'])),
            ('Average Salary', float(data['average_net_salary']))
        ]
        
        for item, value in summary_items:
            worksheet.cell(row=start_row, column=1, value=item).font = Font(bold=True)
            worksheet.cell(row=start_row, column=2, value=value)
            start_row += 1
        
        start_row += 2
        
        # Department breakdown
        if data.get('department_breakdown'):
            worksheet.cell(row=start_row, column=1, value="Department Breakdown").font = Font(bold=True, size=12)
            start_row += 1
            
            headers = ['Department', 'Employees', 'Total Gross', 'Average Salary']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=start_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            start_row += 1
            
            for dept in data['department_breakdown']:
                worksheet.cell(row=start_row, column=1, value=dept.get('department', 'N/A'))
                worksheet.cell(row=start_row, column=2, value=dept.get('employee_count', 0))
                worksheet.cell(row=start_row, column=3, value=float(dept.get('total_gross', 0)))
                worksheet.cell(row=start_row, column=4, value=float(dept.get('avg_salary', 0)))
                start_row += 1
    
    def _add_payroll_summary_to_csv(self, writer):
        """Add payroll summary to CSV"""
        data = self.report_data
        
        writer.writerow(['Payroll Summary'])
        writer.writerow(['Total Employees', data['total_employees']])
        writer.writerow(['Total Gross Salary', data['total_gross_salary']])
        writer.writerow(['Total Net Salary', data['total_net_salary']])
        writer.writerow(['Total Deductions', data['total_deductions']])
        writer.writerow(['Total Bonuses', data['total_bonuses']])
        writer.writerow(['Average Salary', data['average_net_salary']])
        writer.writerow([])
        
        if data.get('department_breakdown'):
            writer.writerow(['Department Breakdown'])
            writer.writerow(['Department', 'Employees', 'Total Gross', 'Average Salary'])
            
            for dept in data['department_breakdown']:
                writer.writerow([
                    dept.get('department', 'N/A'),
                    dept.get('employee_count', 0),
                    dept.get('total_gross', 0),
                    dept.get('avg_salary', 0)
                ])
    
    def _add_performance_to_pdf(self, story, styles):
        """Add performance data to PDF"""
        if not self.report_data:
            story.append(Paragraph("No performance data available for the selected period.", styles['Normal']))
            return
        
        headers = [
            'Employee ID', 'Name', 'Department', 'Attendance Score',
            'Punctuality Score', 'Overall Score'
        ]
        data = [headers]
        
        for employee in self.report_data:
            data.append([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                f"{employee['attendance_reliability_score']}%",
                f"{employee['punctuality_score']}%",
                f"{employee['overall_performance_score']}%"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_performance_to_excel(self, worksheet, start_row, header_font, header_fill, header_alignment):
        """Add performance data to Excel"""
        headers = [
            'Employee ID', 'Name', 'Department', 'Attendance Score',
            'Punctuality Score', 'Overall Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, employee in enumerate(self.report_data, start_row + 1):
            worksheet.cell(row=row_idx, column=1, value=employee['employee_id'])
            worksheet.cell(row=row_idx, column=2, value=employee['employee_name'])
            worksheet.cell(row=row_idx, column=3, value=employee['department'])
            worksheet.cell(row=row_idx, column=4, value=employee['attendance_reliability_score'])
            worksheet.cell(row=row_idx, column=5, value=employee['punctuality_score'])
            worksheet.cell(row=row_idx, column=6, value=employee['overall_performance_score'])
    
    def _add_performance_to_csv(self, writer):
        """Add performance data to CSV"""
        headers = [
            'Employee ID', 'Name', 'Department', 'Attendance Score',
            'Punctuality Score', 'Overall Score'
        ]
        writer.writerow(headers)
        
        for employee in self.report_data:
            writer.writerow([
                employee['employee_id'],
                employee['employee_name'],
                employee['department'],
                employee['attendance_reliability_score'],
                employee['punctuality_score'],
                employee['overall_performance_score']
            ])


def generate_report_filename(report_type, export_format, timestamp=None):
    """Generate a filename for exported reports"""
    if not timestamp:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return f"{report_type}_{timestamp}.{export_format}"


def create_http_response(buffer, filename, content_type):
    """Create HTTP response for file download"""
    response = HttpResponse(buffer.getvalue(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

